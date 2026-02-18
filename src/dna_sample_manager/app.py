"""
DNA Sample Manager - Flask Application for DNA Sample Management
Refactored with Individual model based on TSV import format
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timezone
import os
import secrets
import csv
from io import BytesIO, StringIO


def safe_int(value, default=0, min_val=None, max_val=None):
    """Safely convert a value to int with bounds checking."""
    try:
        result = int(value)
    except (ValueError, TypeError):
        return default
    if min_val is not None:
        result = max(result, min_val)
    if max_val is not None:
        result = min(result, max_val)
    return result

app = Flask(__name__)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))

db = SQLAlchemy()


def create_app(db_path=None):
    """Configure and return the Flask application.

    Args:
        db_path: Absolute path to the SQLite database file.
                 If None, uses 'sqlite:///dna_samples.db' (Flask instance folder).
    """
    if db_path is not None:
        app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    else:
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///dna_samples.db'
    db.init_app(app)
    with app.app_context():
        db.create_all()
    return app

# =============================================================================
# DATABASE MODELS
# =============================================================================

class Individual(db.Model):
    """Model for individuals - based on TSV import format"""
    __tablename__ = 'individual'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    individual_id = db.Column(db.String(50), unique=True, nullable=False, index=True)  # ID from TSV
    aliases = db.Column(db.String(100), nullable=True)
    family_id = db.Column(db.String(50), nullable=True, index=True)  # Index for filtering
    sex = db.Column(db.Integer, nullable=True)  # 0=Unknown, 1=Male, 2=Female
    phenotype = db.Column(db.String(20), nullable=True)
    projects = db.Column(db.String(200), nullable=True)  # Comma-separated list
    other_family_codes = db.Column(db.String(100), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    
    # Relationship to samples
    samples = db.relationship('Sample', backref='individual', lazy=True)

    def get_sex_display(self):
        """Return sex in readable format"""
        if self.sex == 1:
            return 'M'
        elif self.sex == 2:
            return 'F'
        return 'Unknown'

    def get_sample_count(self):
        """Return number of samples for this individual"""
        return db.session.query(db.func.count(Sample.id)).filter(
            Sample.individual_id == self.id
        ).scalar()

    def to_dict(self):
        return {
            'id': self.id,
            'individual_id': self.individual_id,
            'aliases': self.aliases,
            'family_id': self.family_id,
            'sex': self.sex,
            'sex_display': self.get_sex_display(),
            'phenotype': self.phenotype,
            'projects': self.projects,
            'other_family_codes': self.other_family_codes,
            'notes': self.notes,
            'sample_count': self.get_sample_count(),
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class Sample(db.Model):
    """Model for DNA samples - linked to individuals"""
    __tablename__ = 'sample'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    sample_id = db.Column(db.String(20), unique=True, nullable=False, index=True)  # Sample code (e.g., C000A0M)
    individual_id = db.Column(db.Integer, db.ForeignKey('individual.id'), nullable=True, index=True)  # Index for joins
    sample_type = db.Column(db.String(50), nullable=True, index=True)  # DNA, RNA, Blood, etc.
    arrival_date = db.Column(db.Date, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), index=True)  # Index for sorting
    
    # Relationship to tubes
    tubes = db.relationship('Tube', backref='sample', lazy=True)

    def get_tube_count(self):
        """Return number of tubes for this sample"""
        return db.session.query(db.func.count(Tube.id)).filter(
            Tube.sample_id == self.id
        ).scalar()

    def to_dict(self):
        return {
            'id': self.id,
            'sample_code': self.sample_id,  # For template consistency
            'sample_id': self.sample_id,
            'individual_db_id': self.individual_id,  # DB foreign key
            'individual_id': self.individual.individual_id if self.individual else None,  # Individual code
            'family_id': self.individual.family_id if self.individual else None,
            'sample_type': self.sample_type,
            'arrival_date': self.arrival_date.isoformat() if self.arrival_date else None,
            'notes': self.notes,
            'tube_count': self.get_tube_count(),
            'tube_id': self.tubes[0].id if self.tubes else None,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class Box(db.Model):
    """Model for storage boxes"""
    __tablename__ = 'box'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(50), nullable=True)
    box_type = db.Column(db.String(20), default='stock')  # stock or working
    freezer = db.Column(db.String(50), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    
    tubes = db.relationship('Tube', backref='box', lazy=True)

    def get_tube_count(self):
        return db.session.query(db.func.count(Tube.id)).filter(
            Tube.box_id == self.id
        ).scalar()

    def to_dict(self, include_tubes=False):
        """Convert to dict - use include_tubes=True to load tube details"""
        result = {
            'id': self.id,
            'name': self.name,
            'box_type': self.box_type,
            'freezer': self.freezer,
            'notes': self.notes,
            'tube_count': 0  # Default to avoid lazy loading
        }
        
        if include_tubes:
            # Load tubes with their sample info in batch
            from sqlalchemy.orm import joinedload
            tubes_query = Tube.query.filter_by(box_id=self.id)
            tubes = tubes_query.all()
            
            # Get sample and individual data in batch
            sample_ids = [t.sample_id for t in tubes if t.sample_id]
            samples_dict = {}
            if sample_ids:
                samples = db.session.query(
                    Sample.id, 
                    Sample.sample_id,
                    Sample.individual_id
                ).filter(Sample.id.in_(sample_ids)).all()
                
                individual_ids = [s.individual_id for s in samples if s.individual_id]
                individuals_dict = {}
                if individual_ids:
                    individuals = db.session.query(
                        Individual.id,
                        Individual.individual_id
                    ).filter(Individual.id.in_(individual_ids)).all()
                    individuals_dict = {i.id: i.individual_id for i in individuals}
                
                for s in samples:
                    samples_dict[s.id] = {
                        'sample_id': s.sample_id,
                        'individual_id': individuals_dict.get(s.individual_id)
                    }
            
            result['tubes'] = [
                t.to_dict_light(
                    sample_data=samples_dict.get(t.sample_id),
                    box_data={'name': self.name, 'freezer': self.freezer}
                ) for t in tubes
            ]
            result['tube_count'] = len(tubes)
        
        return result


class Tube(db.Model):
    """Model for DNA tubes"""
    __tablename__ = 'tube'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    barcode = db.Column(db.String(20), unique=True, nullable=False, index=True)
    sample_id = db.Column(db.Integer, db.ForeignKey('sample.id'), nullable=True, index=True)
    box_id = db.Column(db.Integer, db.ForeignKey('box.id'), nullable=True, index=True)
    position_row = db.Column(db.Integer, nullable=True)  # 1-8 for A-H
    position_col = db.Column(db.Integer, nullable=True)  # 1-12
    concentration = db.Column(db.Float, nullable=True)  # ng/µL
    quality = db.Column(db.String(20), nullable=True)
    initial_volume = db.Column(db.Float, nullable=True)  # µL
    current_volume = db.Column(db.Float, nullable=True)  # µL
    source = db.Column(db.String(50), nullable=True)  # Blood, Saliva, etc.
    tube_type = db.Column(db.String(20), default='stock')  # stock or working
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    usages = db.relationship('Usage', backref='tube', lazy=True)

    def get_position_display(self):
        """Return position in readable format (e.g., A1, B2, I9)"""
        if self.position_row and self.position_col:
            row = chr(64 + self.position_row) if self.position_row <= 26 else str(self.position_row)
            return f"{row}{self.position_col}"
        return None

    def get_status(self):
        """Determine tube status based on volume"""
        if self.current_volume is None or self.current_volume <= 0:
            return 'Empty'
        elif self.current_volume < 10:
            return 'Critical'
        elif self.initial_volume and self.current_volume < self.initial_volume * 0.25:
            return 'Low'
        return 'Available'

    def to_dict(self):
        return {
            'id': self.id,
            'barcode': self.barcode,
            'sample_id': self.sample_id,
            'sample_code': self.sample.sample_id if self.sample else None,
            'individual_id': self.sample.individual.individual_id if self.sample and self.sample.individual else None,
            'individual_code': self.sample.individual.individual_id if self.sample and self.sample.individual else None,
            'box_id': self.box_id,
            'box_name': self.box.name if self.box else None,
            'freezer': self.box.freezer if self.box else None,
            'position_row': self.position_row,
            'position_col': self.position_col,
            'position_display': self.get_position_display(),
            'concentration': self.concentration,
            'quality': self.quality,
            'initial_volume': self.initial_volume,
            'current_volume': self.current_volume,
            'source': self.source,
            'tube_type': self.tube_type,
            'notes': self.notes,
            'status': self.get_status(),
            'created_at': self.created_at.isoformat() if self.created_at else None
        }
    
    def to_dict_light(self, sample_data=None, box_data=None):
        """Lightweight version that avoids lazy loading - accepts pre-loaded data"""
        return {
            'id': self.id,
            'barcode': self.barcode,
            'sample_id': self.sample_id,
            'sample_code': sample_data.get('sample_id') if sample_data else None,
            'individual_id': sample_data.get('individual_id') if sample_data else None,
            'individual_code': sample_data.get('individual_id') if sample_data else None,
            'box_id': self.box_id,
            'box_name': box_data.get('name') if box_data else None,
            'freezer': box_data.get('freezer') if box_data else None,
            'position_row': self.position_row,
            'position_col': self.position_col,
            'position_display': self.get_position_display(),
            'concentration': self.concentration,
            'quality': self.quality,
            'initial_volume': self.initial_volume,
            'current_volume': self.current_volume,
            'source': self.source,
            'tube_type': self.tube_type,
            'notes': self.notes,
            'status': self.get_status(),
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class Usage(db.Model):
    """Model for tube usage history"""
    __tablename__ = 'usage'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    tube_id = db.Column(db.Integer, db.ForeignKey('tube.id'), nullable=True)
    user_id = db.Column(db.Integer, nullable=True)
    user_name = db.Column(db.String(50), nullable=True)
    date_out = db.Column(db.Date, nullable=True)
    date_return = db.Column(db.Date, nullable=True)
    volume_taken = db.Column(db.Float, nullable=True)
    purpose = db.Column(db.String(100), nullable=True)
    notes = db.Column(db.Text, nullable=True)

    def to_dict(self):
        return {
            'id': self.id,
            'tube_id': self.tube_id,
            'tube_barcode': self.tube.barcode if self.tube else None,
            'user_id': self.user_id,
            'user_name': self.user_name,
            'date_out': self.date_out.isoformat() if self.date_out else None,
            'date_return': self.date_return.isoformat() if self.date_return else None,
            'volume_taken': self.volume_taken,
            'purpose': self.purpose,
            'notes': self.notes
        }


# =============================================================================
# WEB ROUTES
# =============================================================================

@app.route('/')
def index():
    """Dashboard page"""
    return render_template('index_new.html')


@app.route('/individuals')
def individuals_page():
    """Individuals management page"""
    return render_template('individuals.html')


@app.route('/samples')
def samples_page():
    """Samples management page"""
    return render_template('samples.html')


@app.route('/tubes')
def tubes_page():
    """Tubes management page"""
    return render_template('tubes.html')


@app.route('/boxes')
def boxes_page():
    """Boxes management page"""
    return render_template('boxes.html')


@app.route('/history')
def history_page():
    """Usage history page"""
    return render_template('history.html')


# =============================================================================
# API - INDIVIDUALS
# =============================================================================

@app.route('/api/individuals', methods=['GET'])
def get_individuals():
    """Get all individuals with optional filtering"""
    search = request.args.get('search', '')
    family = request.args.get('family', '')
    project = request.args.get('project', '')
    limit = request.args.get('limit', '100')  # Default to 100 for speed
    
    query = Individual.query
    
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            db.or_(
                Individual.individual_id.ilike(search_filter),
                Individual.aliases.ilike(search_filter),
                Individual.family_id.ilike(search_filter),
                Individual.projects.ilike(search_filter)
            )
        )
    
    if family:
        query = query.filter(Individual.family_id == family)
    
    if project:
        query = query.filter(Individual.projects.ilike(f"%{project}%"))
    
    # Always limit to avoid long queries
    max_limit = safe_int(limit, default=100, min_val=1, max_val=1000)
    individuals = query.order_by(Individual.individual_id).limit(max_limit).all()
    
    # Batch load sample counts for these individuals
    from sqlalchemy import func
    ind_ids = [ind.id for ind in individuals]
    sample_counts = {}
    if ind_ids:
        counts = db.session.query(
            Sample.individual_id,
            func.count(Sample.id)
        ).filter(Sample.individual_id.in_(ind_ids)).group_by(Sample.individual_id).all()
        sample_counts = {ind_id: count for ind_id, count in counts}
    
    # Dict conversion with actual sample count
    result = []
    for ind in individuals:
        result.append({
            'id': ind.id,
            'individual_id': ind.individual_id,
            'aliases': ind.aliases,
            'family_id': ind.family_id,
            'sex': ind.sex,
            'sex_display': ind.get_sex_display(),
            'phenotype': ind.phenotype,
            'projects': ind.projects,
            'other_family_codes': ind.other_family_codes,
            'sample_count': sample_counts.get(ind.id, 0)
        })
    
    return jsonify({
        'individuals': result,
        'total': len(result),
        'limited': True
    })


@app.route('/api/individuals/<int:id>', methods=['GET'])
def get_individual(id):
    """Get individual by ID"""
    individual = Individual.query.get_or_404(id)
    return jsonify(individual.to_dict())


@app.route('/api/individuals', methods=['POST'])
def create_individual():
    """Create a new individual"""
    data = request.json
    
    # Check if individual_id already exists
    if Individual.query.filter_by(individual_id=data.get('individual_id')).first():
        return jsonify({'error': 'An individual with this ID already exists'}), 400
    
    individual = Individual(
        individual_id=data.get('individual_id'),
        aliases=data.get('aliases'),
        family_id=data.get('family_id'),
        sex=data.get('sex'),
        phenotype=data.get('phenotype'),
        projects=data.get('projects'),
        other_family_codes=data.get('other_family_codes'),
        notes=data.get('notes')
    )
    
    db.session.add(individual)
    db.session.commit()
    
    return jsonify(individual.to_dict()), 201


@app.route('/api/individuals/<int:id>', methods=['PUT'])
def update_individual(id):
    """Update an individual"""
    individual = Individual.query.get_or_404(id)
    data = request.json
    
    for key in ['individual_id', 'aliases', 'family_id', 'sex', 'phenotype', 
                'projects', 'other_family_codes', 'notes']:
        if key in data:
            setattr(individual, key, data[key])
    
    db.session.commit()
    return jsonify(individual.to_dict())


@app.route('/api/individuals/<int:id>', methods=['DELETE'])
def delete_individual(id):
    """Delete an individual"""
    individual = Individual.query.get_or_404(id)
    
    # Check if individual has samples
    if individual.samples:
        return jsonify({'error': 'Cannot delete individual with associated samples'}), 400
    
    db.session.delete(individual)
    db.session.commit()
    return jsonify({'message': 'Individual deleted successfully'})


# =============================================================================
# API - SAMPLES
# =============================================================================

@app.route('/api/samples', methods=['GET'])
def get_samples():
    """Get all samples with optional filtering and pagination"""
    search = request.args.get('search', '')
    individual = request.args.get('individual', '')
    sample_type = request.args.get('type', '')
    page = safe_int(request.args.get('page', 1), default=1, min_val=1)
    per_page = safe_int(request.args.get('per_page', 100), default=100, min_val=1, max_val=500)
    limit = request.args.get('limit')
    
    # Use joinedload for eager loading
    from sqlalchemy.orm import joinedload
    query = Sample.query.options(joinedload(Sample.individual))
    
    if search:
        search_filter = f"%{search}%"
        query = query.join(Individual, Sample.individual_id == Individual.id, isouter=True).filter(
            db.or_(
                Sample.sample_id.ilike(search_filter),
                Sample.notes.ilike(search_filter),
                Individual.individual_id.ilike(search_filter)
            )
        )
    
    if individual:
        ind_id = safe_int(individual, default=None)
        if ind_id is not None:
            query = query.filter(Sample.individual_id == ind_id)
    
    if sample_type:
        query = query.filter(Sample.sample_type == sample_type)
    
    # If limit is specified, return simple list
    if limit:
        max_limit = safe_int(limit, default=100, min_val=1, max_val=1000)
        samples = query.order_by(Sample.created_at.desc()).limit(max_limit).all()
        return jsonify([s.to_dict() for s in samples])
    
    # Paginated results - skip total count for speed
    samples = query.order_by(Sample.sample_id).offset((page - 1) * per_page).limit(per_page).all()
    
    return jsonify({
        'samples': [s.to_dict() for s in samples],
        'total': len(samples),  # Only count current page for speed
        'page': page,
        'per_page': per_page,
        'pages': 1  # Don't calculate total pages for speed
    })


@app.route('/api/samples/<int:id>', methods=['GET'])
def get_sample(id):
    """Get sample by ID"""
    sample = Sample.query.get_or_404(id)
    return jsonify(sample.to_dict())


@app.route('/api/samples', methods=['POST'])
def create_sample():
    """Create a new sample"""
    data = request.json
    
    # Check if sample_id already exists
    sample_code = data.get('sample_code') or data.get('sample_id')
    if Sample.query.filter_by(sample_id=sample_code).first():
        return jsonify({'error': 'A sample with this ID already exists'}), 400
    
    arrival_date = None
    if data.get('arrival_date'):
        try:
            arrival_date = datetime.strptime(data['arrival_date'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide pour arrival_date. Utilisez AAAA-MM-JJ'}), 400

    sample = Sample(
        sample_id=sample_code,
        individual_id=data.get('individual_id'),
        sample_type=data.get('sample_type'),
        arrival_date=arrival_date,
        notes=data.get('notes')
    )
    
    db.session.add(sample)
    db.session.commit()
    
    return jsonify(sample.to_dict()), 201


@app.route('/api/samples/<int:id>', methods=['PUT'])
def update_sample(id):
    """Update a sample"""
    sample = Sample.query.get_or_404(id)
    data = request.json
    
    # Handle sample_code (alias for sample_id)
    if 'sample_code' in data:
        sample.sample_id = data['sample_code']
    
    for key in ['sample_id', 'individual_id', 'sample_type', 'notes']:
        if key in data:
            setattr(sample, key, data[key])
    
    if 'arrival_date' in data and data['arrival_date']:
        try:
            sample.arrival_date = datetime.strptime(data['arrival_date'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide. Utilisez AAAA-MM-JJ'}), 400
    
    db.session.commit()
    return jsonify(sample.to_dict())


@app.route('/api/samples/<int:id>', methods=['DELETE'])
def delete_sample(id):
    """Delete a sample"""
    sample = Sample.query.get_or_404(id)
    
    if sample.tubes:
        return jsonify({'error': 'Cannot delete sample with associated tubes'}), 400
    
    db.session.delete(sample)
    db.session.commit()
    return jsonify({'message': 'Sample deleted successfully'})


# =============================================================================
# API - TUBES
# =============================================================================

def _build_tubes_query(search='', box='', status='', tube_type='', limit=200):
    """Build a filtered query for tubes. Returns a list of Tube objects."""
    query = Tube.query

    if search:
        search_filter = f"%{search}%"
        matching_sample_ids = db.session.query(Sample.id).filter(
            Sample.sample_id.ilike(search_filter)
        ).all()
        matching_sample_ids = [s[0] for s in matching_sample_ids]

        matching_individuals = db.session.query(Individual.id).filter(
            Individual.individual_id.ilike(search_filter)
        ).all()
        matching_individual_ids = [i[0] for i in matching_individuals]

        if matching_individual_ids:
            matching_samples_from_individuals = db.session.query(Sample.id).filter(
                Sample.individual_id.in_(matching_individual_ids)
            ).all()
            matching_sample_ids.extend([s[0] for s in matching_samples_from_individuals])

        query = query.filter(
            db.or_(
                Tube.barcode.ilike(search_filter),
                Tube.sample_id.in_(matching_sample_ids) if matching_sample_ids else False
            )
        )

    if tube_type:
        query = query.filter(Tube.tube_type == tube_type)

    if status:
        if status == 'Empty':
            query = query.filter(db.or_(Tube.current_volume == None, Tube.current_volume <= 0))
        elif status == 'Critical':
            query = query.filter(Tube.current_volume > 0, Tube.current_volume < 10)
        elif status == 'Low':
            query = query.filter(
                Tube.current_volume >= 10,
                Tube.initial_volume != None,
                Tube.current_volume < Tube.initial_volume * 0.25
            )
        elif status == 'Available':
            query = query.filter(
                Tube.current_volume >= 10,
                db.or_(
                    Tube.initial_volume == None,
                    Tube.current_volume >= Tube.initial_volume * 0.25
                )
            )

    if box:
        query = query.filter(Tube.box_id == safe_int(box, default=0))
        return query.order_by(Tube.barcode).all()
    else:
        return query.order_by(Tube.barcode).limit(limit).all()


def _tubes_to_dicts(tubes):
    """Convert a list of Tube ORM objects to dicts with batch-loaded related data."""
    sample_ids = [t.sample_id for t in tubes if t.sample_id]
    box_ids = [t.box_id for t in tubes if t.box_id]

    samples_dict = {}
    if sample_ids:
        samples = db.session.query(
            Sample.id, Sample.sample_id, Sample.individual_id
        ).filter(Sample.id.in_(sample_ids)).all()

        individual_ids = [s.individual_id for s in samples if s.individual_id]
        individuals_dict = {}
        if individual_ids:
            individuals = db.session.query(
                Individual.id, Individual.individual_id
            ).filter(Individual.id.in_(individual_ids)).all()
            individuals_dict = {i.id: i.individual_id for i in individuals}

        for s in samples:
            samples_dict[s.id] = {
                'sample_id': s.sample_id,
                'individual_id': individuals_dict.get(s.individual_id)
            }

    boxes_dict = {}
    if box_ids:
        boxes_q = db.session.query(
            Box.id, Box.name, Box.freezer
        ).filter(Box.id.in_(box_ids)).all()
        boxes_dict = {b.id: {'name': b.name, 'freezer': b.freezer} for b in boxes_q}

    result = []
    for t in tubes:
        sample_data = samples_dict.get(t.sample_id, {})
        box_data = boxes_dict.get(t.box_id, {})
        result.append({
            'id': t.id,
            'barcode': t.barcode,
            'sample_id': t.sample_id,
            'sample_code': sample_data.get('sample_id'),
            'individual_id': sample_data.get('individual_id'),
            'individual_code': sample_data.get('individual_id'),
            'box_id': t.box_id,
            'box_name': box_data.get('name') if box_data else None,
            'freezer': box_data.get('freezer') if box_data else None,
            'position_row': t.position_row,
            'position_col': t.position_col,
            'position_display': t.get_position_display(),
            'concentration': t.concentration,
            'quality': t.quality,
            'initial_volume': t.initial_volume,
            'current_volume': t.current_volume,
            'source': t.source,
            'tube_type': t.tube_type,
            'notes': t.notes,
            'status': t.get_status(),
            'created_at': t.created_at.isoformat() if t.created_at else None
        })
    return result


@app.route('/api/tubes', methods=['GET'])
def get_tubes():
    """Get all tubes with search in barcode, sample_id, and individual_id"""
    try:
        search = request.args.get('search', '')
        box = request.args.get('box', '')
        status = request.args.get('status', '')
        tube_type = request.args.get('type', '')
        limit = safe_int(request.args.get('limit', 200), default=200, min_val=1, max_val=5000)

        tubes = _build_tubes_query(search=search, box=box, status=status,
                                   tube_type=tube_type, limit=limit)
        result = _tubes_to_dicts(tubes)
        return jsonify(result)
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Export columns: (dict_key, french_header)
EXPORT_COLUMNS = [
    ('barcode', 'Barcode'),
    ('sample_code', 'Code Échantillon'),
    ('individual_id', 'ID Individu'),
    ('box_name', 'Boîte'),
    ('freezer', 'Congélateur'),
    ('position_display', 'Position'),
    ('concentration', 'Concentration (ng/µL)'),
    ('quality', 'Qualité'),
    ('initial_volume', 'Volume Initial (µL)'),
    ('current_volume', 'Volume Actuel (µL)'),
    ('source', 'Source'),
    ('tube_type', 'Type'),
    ('status', 'Statut'),
    ('notes', 'Notes'),
]


@app.route('/api/tubes/export', methods=['GET'])
def export_tubes():
    """Export filtered tubes as TSV or Excel file."""
    try:
        fmt = request.args.get('format', 'tsv')
        if fmt not in ('tsv', 'xlsx'):
            return jsonify({'error': 'Format must be tsv or xlsx'}), 400

        search = request.args.get('search', '')
        box = request.args.get('box', '')
        status = request.args.get('status', '')
        tube_type = request.args.get('type', '')

        tubes = _build_tubes_query(search=search, box=box, status=status,
                                   tube_type=tube_type, limit=50000)
        rows = _tubes_to_dicts(tubes)

        today = datetime.now().strftime('%Y%m%d')

        if fmt == 'tsv':
            text_output = StringIO()
            writer = csv.writer(text_output, delimiter='\t')
            writer.writerow([col[1] for col in EXPORT_COLUMNS])
            for row in rows:
                writer.writerow([row.get(col[0], '') or '' for col in EXPORT_COLUMNS])

            output = BytesIO()
            output.write(text_output.getvalue().encode('utf-8-sig'))
            output.seek(0)

            return send_file(
                output,
                mimetype='text/tab-separated-values',
                as_attachment=True,
                download_name=f'tubes_export_{today}.tsv'
            )

        else:  # xlsx
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = 'Tubes'

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2D3250', end_color='2D3250', fill_type='solid')

            headers = [col[1] for col in EXPORT_COLUMNS]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row in rows:
                ws.append([row.get(col[0], '') or '' for col in EXPORT_COLUMNS])

            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col_idx, column in enumerate(EXPORT_COLUMNS, 1):
                max_length = len(column[1])
                for row_data in rows[:100]:
                    val = str(row_data.get(column[0], '') or '')
                    max_length = max(max_length, len(val))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 35)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'tubes_export_{today}.xlsx'
            )

    except Exception as e:
        print(f"EXPORT ERROR: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/tubes/<int:id>', methods=['GET'])
def get_tube(id):
    """Get tube by ID"""
    tube = Tube.query.get_or_404(id)
    return jsonify(tube.to_dict())


@app.route('/api/tubes', methods=['POST'])
def create_tube():
    """Create a new tube"""
    data = request.json
    
    # Check if barcode already exists
    if Tube.query.filter_by(barcode=data.get('barcode')).first():
        return jsonify({'error': 'A tube with this barcode already exists'}), 400
    
    tube = Tube(
        barcode=data.get('barcode'),
        sample_id=data.get('sample_id'),
        box_id=data.get('box_id'),
        position_row=data.get('position_row'),
        position_col=data.get('position_col'),
        concentration=data.get('concentration'),
        quality=data.get('quality'),
        initial_volume=data.get('initial_volume'),
        current_volume=data.get('current_volume'),
        source=data.get('source'),
        tube_type=data.get('tube_type', 'stock'),
        notes=data.get('notes')
    )
    
    db.session.add(tube)
    db.session.commit()
    
    return jsonify(tube.to_dict()), 201


@app.route('/api/tubes/<int:id>', methods=['PUT'])
def update_tube(id):
    """Update a tube"""
    tube = Tube.query.get_or_404(id)
    data = request.json
    
    for key in ['barcode', 'sample_id', 'box_id', 'position_row', 'position_col',
                'concentration', 'quality', 'initial_volume', 'current_volume',
                'source', 'tube_type', 'notes']:
        if key in data:
            setattr(tube, key, data[key])
    
    db.session.commit()
    return jsonify(tube.to_dict())


@app.route('/api/tubes/<int:id>', methods=['DELETE'])
def delete_tube(id):
    """Delete a tube"""
    tube = Tube.query.get_or_404(id)

    if tube.usages:
        return jsonify({'error': 'Cannot delete tube with usage history. Clear usage records first.'}), 400

    db.session.delete(tube)
    db.session.commit()
    return jsonify({'message': 'Tube deleted successfully'})


# =============================================================================
# API - BOXES
# =============================================================================

@app.route('/api/boxes', methods=['GET'])
def get_boxes():
    """Get all boxes with tube counts"""
    boxes = Box.query.order_by(Box.name).all()
    
    # Get tube counts in one query
    box_ids = [b.id for b in boxes]
    tube_counts = {}
    if box_ids:
        from sqlalchemy import func
        counts = db.session.query(
            Tube.box_id,
            func.count(Tube.id)
        ).filter(Tube.box_id.in_(box_ids)).group_by(Tube.box_id).all()
        tube_counts = {box_id: count for box_id, count in counts}
    
    # Build response
    result = []
    for b in boxes:
        box_dict = b.to_dict()
        box_dict['tube_count'] = tube_counts.get(b.id, 0)
        result.append(box_dict)
    
    return jsonify(result)


@app.route('/api/boxes/<int:id>', methods=['GET'])
def get_box(id):
    """Get box by ID with tubes"""
    box = Box.query.get_or_404(id)
    return jsonify(box.to_dict(include_tubes=True))


@app.route('/api/boxes', methods=['POST'])
def create_box():
    """Create a new box"""
    data = request.json
    
    box = Box(
        name=data.get('name'),
        box_type=data.get('box_type', 'stock'),
        freezer=data.get('freezer'),
        notes=data.get('notes')
    )
    
    db.session.add(box)
    db.session.commit()
    
    return jsonify(box.to_dict()), 201


@app.route('/api/boxes/<int:id>', methods=['PUT'])
def update_box(id):
    """Update a box"""
    box = Box.query.get_or_404(id)
    data = request.json
    
    for key in ['name', 'box_type', 'freezer', 'notes']:
        if key in data:
            setattr(box, key, data[key])
    
    db.session.commit()
    return jsonify(box.to_dict())


@app.route('/api/boxes/<int:id>', methods=['DELETE'])
def delete_box(id):
    """Delete a box"""
    box = Box.query.get_or_404(id)
    
    if box.tubes:
        return jsonify({'error': 'Cannot delete box containing tubes'}), 400
    
    db.session.delete(box)
    db.session.commit()
    return jsonify({'message': 'Box deleted successfully'})


# =============================================================================
# API - USAGE HISTORY
# =============================================================================

@app.route('/api/usages', methods=['GET'])
def get_usages():
    """Get usage history"""
    tube = request.args.get('tube', '')
    
    query = Usage.query
    
    if tube:
        tube_id = safe_int(tube, default=None)
        if tube_id is not None:
            query = query.filter(Usage.tube_id == tube_id)
    
    usages = query.order_by(Usage.date_out.desc()).limit(500).all()
    return jsonify([u.to_dict() for u in usages])


@app.route('/api/usages', methods=['POST'])
def create_usage():
    """Create a new usage record"""
    data = request.json

    # Validate volume_taken
    volume_taken = None
    if data.get('volume_taken') is not None:
        try:
            volume_taken = float(data['volume_taken'])
        except (ValueError, TypeError):
            return jsonify({'error': 'volume_taken doit être un nombre'}), 400
        if volume_taken <= 0:
            return jsonify({'error': 'volume_taken doit être positif'}), 400

    # Validate tube exists if tube_id provided
    tube = None
    if data.get('tube_id'):
        tube = Tube.query.get(data['tube_id'])
        if not tube:
            return jsonify({'error': 'Tube non trouvé'}), 404

    # Check sufficient volume
    if tube and volume_taken and tube.current_volume is not None:
        if volume_taken > tube.current_volume:
            return jsonify({
                'error': f'Volume insuffisant. Disponible : {tube.current_volume} µL, demandé : {volume_taken} µL'
            }), 400

    # Parse date safely
    date_out = datetime.now().date()
    if data.get('date_out'):
        try:
            date_out = datetime.strptime(data['date_out'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Format de date invalide. Utilisez AAAA-MM-JJ'}), 400

    usage = Usage(
        tube_id=data.get('tube_id'),
        user_name=data.get('user_name'),
        date_out=date_out,
        volume_taken=volume_taken,
        purpose=data.get('purpose'),
        notes=data.get('notes')
    )

    # Update tube current volume
    if tube and volume_taken and tube.current_volume is not None:
        tube.current_volume -= volume_taken

    db.session.add(usage)
    db.session.commit()

    return jsonify(usage.to_dict()), 201


# =============================================================================
# API - STATISTICS
# =============================================================================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get dashboard statistics - optimized"""
    # Use scalar queries for counts (faster)
    total_individuals = db.session.query(db.func.count(Individual.id)).scalar()
    total_samples = db.session.query(db.func.count(Sample.id)).scalar()
    total_tubes = db.session.query(db.func.count(Tube.id)).scalar()
    total_boxes = db.session.query(db.func.count(Box.id)).scalar()
    
    # Top families - already optimized
    family_counts = db.session.query(
        Individual.family_id, db.func.count(Individual.id)
    ).filter(Individual.family_id != None, Individual.family_id != '').group_by(
        Individual.family_id
    ).order_by(db.func.count(Individual.id).desc()).limit(5).all()
    
    # Top projects - load only the projects column for all individuals
    proj_rows = db.session.query(Individual.projects).filter(
        Individual.projects != None, Individual.projects != ''
    ).all()
    project_counts = {}
    for (proj_str,) in proj_rows:
        if proj_str:
            for proj in proj_str.split(','):
                proj = proj.strip()
                if proj:
                    project_counts[proj] = project_counts.get(proj, 0) + 1
    top_projects = sorted(project_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return jsonify({
        'individuals': total_individuals,
        'samples': total_samples,
        'tubes': total_tubes,
        'boxes': total_boxes,
        'top_families': [{'family_id': f[0], 'count': f[1]} for f in family_counts],
        'top_projects': [{'project': p[0], 'count': p[1]} for p in top_projects]
    })


@app.route('/api/families', methods=['GET'])
def get_families():
    """Get list of unique families"""
    families = db.session.query(Individual.family_id).filter(
        Individual.family_id != None, Individual.family_id != ''
    ).distinct().order_by(Individual.family_id).all()
    return jsonify([f[0] for f in families])


@app.route('/api/projects', methods=['GET'])
def get_projects():
    """Get list of unique projects - optimized to load only the projects column"""
    rows = db.session.query(Individual.projects).filter(
        Individual.projects != None, Individual.projects != ''
    ).all()
    projects = set()
    for (proj_str,) in rows:
        if proj_str:
            for proj in proj_str.split(','):
                p = proj.strip()
                if p:
                    projects.add(p)
    return jsonify(sorted(projects))


@app.route('/api/samples/types', methods=['GET'])
def get_sample_types():
    """Get list of unique sample types"""
    types = db.session.query(Sample.sample_type).filter(
        Sample.sample_type != None, Sample.sample_type != ''
    ).distinct().order_by(Sample.sample_type).all()
    return jsonify({'types': [t[0] for t in types]})


# =============================================================================
# API - SUJETS (Fusion Individuals + Samples)
# =============================================================================

@app.route('/sujets')
def sujets_page():
    """Page de gestion des sujets (Individuals + Samples)"""
    return render_template('sujets.html')


@app.route('/api/sujets', methods=['GET'])
def get_sujets():
    """Get all individuals - paginated and ultra-optimized for large datasets"""
    search = request.args.get('search', '')
    family = request.args.get('family', '')
    project = request.args.get('project', '')
    page = safe_int(request.args.get('page', 1), default=1, min_val=1)
    per_page = safe_int(request.args.get('per_page', 25), default=25, min_val=1, max_val=100)
    
    # Build base query
    query = Individual.query
    
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            db.or_(
                Individual.individual_id.ilike(search_filter),
                Individual.aliases.ilike(search_filter),
                Individual.family_id.ilike(search_filter),
                Individual.projects.ilike(search_filter)
            )
        )
    
    if family:
        query = query.filter(Individual.family_id == family)
    
    if project:
        query = query.filter(Individual.projects.ilike(f"%{project}%"))
    
    # Count efficiently - only when filtering
    if search or family or project:
        total = query.count()
    else:
        # For full table, use fast query
        total = db.session.query(Individual.id).count()
    
    # Get paginated results
    individuals = query.order_by(Individual.individual_id).offset((page - 1) * per_page).limit(per_page).all()
    
    total_pages = max(1, (total + per_page - 1) // per_page)
    
    # Get tube counts per individual (via samples) in one query
    individual_ids = [ind.id for ind in individuals]
    tube_counts = {}
    if individual_ids:
        from sqlalchemy import func, text
        counts = db.session.query(
            Sample.individual_id,
            func.count(Tube.id)
        ).join(Tube, Tube.sample_id == Sample.id
        ).filter(Sample.individual_id.in_(individual_ids)
        ).group_by(Sample.individual_id).all()
        tube_counts = {ind_id: count for ind_id, count in counts}
    
    # Build result with tube counts
    result = []
    for ind in individuals:
        result.append({
            'id': ind.id,
            'individual_id': ind.individual_id,
            'aliases': ind.aliases,
            'family_id': ind.family_id,
            'sex': ind.sex,
            'sex_display': ind.get_sex_display(),
            'phenotype': ind.phenotype,
            'projects': ind.projects,
            'other_family_codes': ind.other_family_codes,
            'notes': ind.notes,
            'sample_count': tube_counts.get(ind.id, 0),
            'samples': [],  # Load on demand via details
            'created_at': ind.created_at.isoformat() if ind.created_at else None,
            'updated_at': ind.updated_at.isoformat() if ind.updated_at else None
        })
    
    return jsonify({
        'sujets': result,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': total_pages
    })


@app.route('/api/sujets/<int:id>', methods=['GET'])
def get_sujet(id):
    """Get individual by ID with samples"""
    individual = Individual.query.get_or_404(id)
    data = {
        'id': individual.id,
        'individual_id': individual.individual_id,
        'aliases': individual.aliases,
        'family_id': individual.family_id,
        'sex': individual.sex,
        'sex_display': individual.get_sex_display(),
        'phenotype': individual.phenotype,
        'projects': individual.projects,
        'other_family_codes': individual.other_family_codes,
        'notes': individual.notes,
        'created_at': individual.created_at.isoformat() if individual.created_at else None,
        'updated_at': individual.updated_at.isoformat() if individual.updated_at else None
    }
    
    # Load samples via raw SQL to avoid date parsing issues
    from sqlalchemy import func, text
    sample_rows = db.session.execute(
        text("SELECT id, sample_id, sample_type FROM sample WHERE individual_id = :ind_id"),
        {'ind_id': individual.id}
    ).fetchall()
    
    samples_list = []
    for row in sample_rows:
        s_id, s_sample_id, s_type = row[0], row[1], row[2]
        # Get tubes for this sample with box info
        tube_rows = db.session.execute(
            text("""SELECT t.id, t.barcode, t.position_row, t.position_col, b.name, b.freezer
                    FROM tube t LEFT JOIN box b ON t.box_id = b.id
                    WHERE t.sample_id = :sid
                    ORDER BY b.name, t.position_row, t.position_col"""),
            {'sid': s_id}
        ).fetchall()
        tubes_list = []
        for tr in tube_rows:
            # Build position display
            pos_display = ''
            if tr[2] and tr[3]:
                row_letter = chr(ord('A') + tr[2] - 1) if 1 <= tr[2] <= 26 else str(tr[2])
                pos_display = f"{row_letter}{tr[3]}"
            tubes_list.append({
                'id': tr[0],
                'barcode': tr[1],
                'position_display': pos_display,
                'box_name': tr[4],
                'freezer': tr[5]
            })
        samples_list.append({
            'id': s_id,
            'sample_id': s_sample_id,
            'sample_type': s_type,
            'tube_count': len(tubes_list),
            'tubes': tubes_list
        })
    
    data['samples'] = samples_list
    data['sample_count'] = len(samples_list)
    # Total tube count
    data['tube_count'] = sum(s['tube_count'] for s in samples_list)
    return jsonify(data)


@app.route('/api/sujets', methods=['POST'])
def create_sujet():
    """Create a new individual (subject)"""
    data = request.json
    
    if Individual.query.filter_by(individual_id=data.get('individual_id')).first():
        return jsonify({'error': 'Un sujet avec cet ID existe déjà'}), 400
    
    individual = Individual(
        individual_id=data.get('individual_id'),
        aliases=data.get('aliases'),
        family_id=data.get('family_id'),
        sex=data.get('sex'),
        phenotype=data.get('phenotype'),
        projects=data.get('projects'),
        other_family_codes=data.get('other_family_codes'),
        notes=data.get('notes')
    )
    
    db.session.add(individual)
    db.session.commit()
    
    return jsonify(individual.to_dict()), 201


@app.route('/api/sujets/<int:id>', methods=['PUT'])
def update_sujet(id):
    """Update an individual"""
    individual = Individual.query.get_or_404(id)
    data = request.json
    
    for key in ['individual_id', 'aliases', 'family_id', 'sex', 'phenotype', 
                'projects', 'other_family_codes', 'notes']:
        if key in data:
            setattr(individual, key, data[key])
    
    db.session.commit()
    return jsonify(individual.to_dict())


@app.route('/api/sujets/<int:id>', methods=['DELETE'])
def delete_sujet(id):
    """Delete an individual"""
    individual = Individual.query.get_or_404(id)
    
    if individual.samples:
        return jsonify({'error': 'Impossible de supprimer un sujet avec des échantillons associés'}), 400
    
    db.session.delete(individual)
    db.session.commit()
    return jsonify({'message': 'Sujet supprimé avec succès'})


@app.route('/api/sujets/<int:id>/samples', methods=['POST'])
def add_sample_to_sujet(id):
    """Add a sample to an individual"""
    individual = Individual.query.get_or_404(id)
    data = request.json
    
    sample_code = data.get('sample_id')
    if Sample.query.filter_by(sample_id=sample_code).first():
        return jsonify({'error': 'Un échantillon avec ce code existe déjà'}), 400
    
    sample = Sample(
        sample_id=sample_code,
        individual_id=individual.id,
        sample_type=data.get('sample_type'),
        notes=data.get('notes')
    )
    
    db.session.add(sample)
    db.session.commit()
    
    return jsonify({'message': 'Échantillon ajouté', 'sample': {
        'id': sample.id,
        'sample_id': sample.sample_id,
        'sample_type': sample.sample_type
    }}), 201


if __name__ == '__main__':
    flask_app = create_app()
    print("=" * 80)
    print("Server starting on http://127.0.0.1:5003")
    print("=" * 80)
    flask_app.run(debug=False, port=5003, host='127.0.0.1', threaded=True)
